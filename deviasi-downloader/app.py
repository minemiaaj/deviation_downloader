import os
import time
import uuid
import threading
from datetime import datetime
from io import BytesIO
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, jsonify, send_file

# ====== KONFIGURASI ======
LOGIN_URL = "https://epengendalian.sultraprov.go.id/auth/login"
DOWNLOAD_URL_TEMPLATE = "https://epengendalian.sultraprov.go.id/modules/super-admin/apbd.php?opd_id={opd_id}&tahun={tahun}&bulan={bulan}&export=excel"
USERNAME = "SA_Irvhan"
PASSWORD = "123"

OPD_IDS = [
    4,5,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,
    31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55
]

app = Flask(__name__)
tasks = {}

# ====== FUNGSI UTAMA ======
def setup_driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-gpu")
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=chrome_options
    )
    return driver

def get_session_from_selenium():
    print("🔐 Login (headless) ...")
    driver = setup_driver()
    wait = WebDriverWait(driver, 20)
    try:
        driver.get(LOGIN_URL)
        wait.until(EC.presence_of_element_located((By.NAME, "username"))).send_keys(USERNAME)
        driver.find_element(By.NAME, "password").send_keys(PASSWORD)
        driver.find_element(By.XPATH, "//button[@type='submit']").click()
        wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Dashboard')]")))
        print("✅ Login berhasil.\n")
        selenium_cookies = driver.get_cookies()
        driver.quit()
        session = requests.Session()
        for cookie in selenium_cookies:
            session.cookies.set(cookie['name'], cookie['value'])
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        return session
    except Exception as e:
        print(f"❌ Login gagal: {e}")
        driver.quit()
        return None

def get_workbook(session, opd_id, tahun, bulan):
    url = DOWNLOAD_URL_TEMPLATE.format(opd_id=opd_id, tahun=tahun, bulan=bulan)
    try:
        response = session.get(url, timeout=60)
        if response.status_code == 200:
            return BytesIO(response.content)
        else:
            return None
    except Exception:
        return None

def parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(',', '.')
        if cleaned == '' or cleaned == '-':
            return 0.0
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None

def ekstrak_info(file_stream, opd_id):
    """
    Cari TOTAL di kolom B dulu, jika tidak ada baru kolom A.
    Ambil N (13) dan D (3). Hitung deviasi = (ΣN/ΣD)*100%.
    """
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        sheets = wb.sheetnames

        # Nama OPD
        nama_opd = None
        ws0 = wb[sheets[0]]
        for row in ws0.iter_rows(min_row=1, max_row=10, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and "opd:" in cell.lower():
                    idx = cell.lower().find("opd:")
                    nama_opd = cell[idx+4:].strip()
                    if nama_opd.startswith(':'):
                        nama_opd = nama_opd[1:].strip()
                    break
            if nama_opd:
                break
        if not nama_opd:
            for row in ws0.iter_rows(min_row=1, max_row=10, values_only=True):
                if row[0] and isinstance(row[0], str) and row[0].strip().upper() == "OPD":
                    if len(row) > 1 and row[1]:
                        nama_opd = str(row[1]).strip()
                        break
        if not nama_opd:
            nama_opd = f"OPD_ID_{opd_id}"

        total_N = 0.0
        total_D = 0.0
        for sheet_name in sheets:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                # Cek kolom B (1)
                col_b = row[1] if len(row) > 1 else None
                is_total = False
                if col_b and isinstance(col_b, str) and col_b.strip().upper() == "TOTAL":
                    is_total = True
                else:
                    # Cek kolom A (0)
                    col_a = row[0] if len(row) > 0 else None
                    if col_a and isinstance(col_a, str) and col_a.strip().upper() == "TOTAL":
                        is_total = True
                if is_total:
                    n_val = parse_number(row[13]) if len(row) > 13 else None
                    d_val = parse_number(row[3])  if len(row) > 3  else None
                    if n_val is not None and d_val is not None:
                        total_N += n_val
                        total_D += d_val
                    break
        wb.close()

        if total_D != 0:
            deviasi = (total_N / total_D) * 100
            return nama_opd, deviasi
        else:
            return nama_opd, None

    except Exception as e:
        print(f"   ❌ Error parsing OPD {opd_id}: {e}")
        return None, None

def generate_excel(hasil):
    """Buat file Excel modern tanpa OPD ID, urut nama."""
    hasil_sorted = sorted(hasil, key=lambda x: (x['nama'] or '').lower())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deviasi OPD"

    headers = ["No", "Nama OPD", "Deviasi (%)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    for i, item in enumerate(hasil_sorted, start=1):
        nama = item['nama'] if item['nama'] else f"OPD_ID_{item['opd_id']}"
        deviasi = item['deviasi']
        dev_str = f"{deviasi:.2f}" if deviasi is not None else "N/A"
        row_data = [i, nama, dev_str]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i+1, column=col, value=value)
            cell.font = Font(size=11)
            cell.border = thin_border
            if col == 1:
                cell.alignment = center_align
            elif col == 2:
                cell.alignment = left_align
            elif col == 3:
                cell.alignment = right_align
                cell.number_format = '0.00'

    for col in range(1, 4):
        max_length = 0
        for row in range(1, len(hasil_sorted)+2):
            val = str(ws.cell(row=row, column=col).value)
            length = max(len(line) for line in val.split('\n')) * 1.2
            if length > max_length:
                max_length = length
        adjusted = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[get_column_letter(col)].width = adjusted

    ws.freeze_panes = "A2"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def proses_semua_opd(bulan, tahun, progress_callback=None):
    session = get_session_from_selenium()
    if not session:
        return None

    hasil = []
    total = len(OPD_IDS)
    for i, opd_id in enumerate(OPD_IDS, 1):
        stream = get_workbook(session, opd_id, tahun, bulan)
        nama, deviasi = None, None
        if stream:
            nama, deviasi = ekstrak_info(stream, opd_id)
        else:
            print(f"   ❌ Gagal mengunduh OPD {opd_id}")

        hasil.append({
            "opd_id": opd_id,
            "nama": nama if nama else f"OPD_ID_{opd_id}",
            "deviasi": deviasi
        })

        if progress_callback:
            progress_callback({
                "selesai": i,
                "total": total,
                "opd_id": opd_id,
                "nama": nama,
                "deviasi": deviasi
            })
        time.sleep(0.3)
    return hasil

# ====== ROUTE FLASK ======
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/start', methods=['POST'])
def start():
    data = request.get_json()
    bulan = int(data.get('bulan', 7))
    tahun = int(data.get('tahun', 2026))
    task_id = str(uuid.uuid4())
    tasks[task_id] = {
        "status": "starting",
        "progress": {"selesai": 0, "total": len(OPD_IDS)}
    }

    def run():
        def cb(info):
            tasks[task_id]["progress"] = info
            tasks[task_id]["status"] = "running"
        try:
            hasil = proses_semua_opd(bulan, tahun, progress_callback=cb)
            tasks[task_id]["status"] = "completed"
            tasks[task_id]["hasil"] = hasil
        except Exception as e:
            tasks[task_id]["status"] = "error"
            tasks[task_id]["error"] = str(e)

    thread = threading.Thread(target=run)
    thread.start()
    return jsonify({"task_id": task_id})

@app.route('/status/<task_id>')
def task_status(task_id):
    task = tasks.get(task_id)
    if not task:
        return jsonify({"error": "Task tidak ditemukan"}), 404
    return jsonify(task)

@app.route('/download_excel/<task_id>')
def download_excel(task_id):
    task = tasks.get(task_id)
    if not task or task.get('status') != 'completed':
        return "File tidak tersedia", 404
    output = generate_excel(task['hasil'])
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"deviasi_opd_{timestamp}.xlsx"
    )

if __name__ == '__main__':
    app.run(debug=True, threaded=True)